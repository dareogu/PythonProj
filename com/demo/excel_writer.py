#!/usr/bin/env python
# -*- encoding: utf-8 -*-
import warnings
from openpyxl import load_workbook

warnings.simplefilter(action='ignore', category=DeprecationWarning)  #忽略DeprecationWarning

book = load_workbook("Account.xlsx")  # 得到Excel文件的book对象，实例化对象

sheet0 = book.worksheets[0]  # 通过sheet索引获得sheet对象

print("1、", sheet0)

sheet_name = book.get_sheet_names()[1]  # 获得指定索引的sheet表名字

print("2、", sheet_name)

sheet1 = book.get_sheet_by_name(sheet_name)  # 通过sheet名字来获取，当然如果知道sheet名字就可以直接指定

nrows = sheet0.max_row  # 获取行总数

print("3、", nrows)

# 循环打印每一行的内容

for row in sheet1.iter_rows(values_only=True):
    print(row)

ncols = sheet0.max_column  # 获取列总数

print("4、", ncols)

for row in sheet0.iter_rows(min_row=1, max_row=1, values_only=True): # 获得第1行的数据列表
    print(row)

for row in sheet0.iter_rows(min_col=1, max_col=1, values_only=True): # 获得第1列的数据列表
    col_data = row[0]  #可去除列值后面跟着的逗号
    print(col_data)

# 通过坐标读取表格中的数据

cell_value1 = sheet0.cell(1, 1).value

print("6、", cell_value1)

cell_value2 = sheet0.cell(2, 2).value

print("7、", cell_value2)
